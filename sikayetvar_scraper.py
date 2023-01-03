# -*- coding: utf-8 -*-
"""
Created on Fri Dec  2 19:42:16 2022

@author: yakou
"""


from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

header = {"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36"}
# headerdaki user agenti kendi user agentiniz olarak değiştiriniz.


brand = input("Marka giriniz ... : ")
brand = brand.replace(" ","-")
file = input("Oluşturulacak Excel dosyasının adı ... : ")
first_page = input("Hangi sayfadan başlayalım? ... :")
end_page = int(input("Hangi sayfada bitirelim? ... :"))
end_page = end_page +1


say = 1
sabit = 1



for i in range(int(first_page),end_page):
    first_page=int(first_page)
    page1 = "https://www.sikayetvar.com/{}?page=".format(brand)
    page2 = page1+str(first_page)
    first_page+=1
    r = requests.get(page2,headers=header)
    soup = BeautifulSoup(r.content, "lxml")
    complaints = soup.find_all("article",attrs={"class":"story-card"})
    for complaint in complaints:
        
        
        link = complaint.find("h2",attrs={"class":"complaint-title"})
        skip = complaint.find("div",attrs={"class":"story-success active"})
    
        if link == skip:
            continue
         
    
        try:
            link_continue = link.a.get("href")
        except:
            continue
    
        
        link_first = "https://www.sikayetvar.com"
        link_complate =link_first + link_continue
        


        
        complaint_ = requests.get(link_complate,headers = header)
        complaint_soup = BeautifulSoup(complaint_.content, "lxml")
    
        x_details = complaint_soup.find("div",attrs={"class":"card-text"})
        
        link1 = complaint_soup.find("span",attrs={"class":"username"})
        link2 = link1.span.get("title")
        
        #Writing the excel
        ws.cell(row=say, column=sabit).value = link2
        ws.cell(row=say, column=sabit+1).value = link_complate
        ws.cell(row=say, column=sabit+2).value = x_details.text
        say+=1
        wb.save("{}.xlsx".format(file))
        
        
        
        print(link2)
        print(link_complate)
        print(x_details.text)
        print("****************")
        
        
#♠coded by yakup
